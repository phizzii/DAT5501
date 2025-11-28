import xlsxwriter
from openpyxl import load_workbook
import pandas as pd
import os.path
import os
import shutil

# function to search for a folder in a directory
def searchFolder(folder_name):
    print("Finding...")
    
    # using depth first tree traversal to look in each folder for folder name specified
    for root, dirs, files in os.walk('/Users'):
        
        # if found, folder path is joined together with root and returned
        if folder_name in dirs:
            print("Found")
            path = os.path.join(root, folder_name)
            print(path)
            return path
    
    # if not found, None is returned
    return None

# function to search for a file in a directory
def searchFile(file_name):
    print("Finding...")
    
    # using depth first tree traversal to look in each folder for file name specified
    for root, dirs, files in os.walk('/Users'):
        
        # if found, file path is joined with root and returned
        if file_name in files:
            print("Found")
            path = os.path.join(root, file_name)
            print(path)
            return path
    
    # if not found, None is returned
    return None

# function to delete a non empty directory / folder

def delNonEmptyDir(folder_path):
    
    # using shutil module imported to remove 'tree' (folder specified from folder path)
    shutil.rmtree(folder_path)
    print('Folder and its content removed')
    
# function to create directory (folder)
def createDir():
    dirName = input('What would you like the new directory to be called?')
    
    # if name inputted matches directory that exists, user is taken back to choose another folder name
    if os.path.isdir(dirName) == True:
        print('Please choose a different name - this directory already exists')
        createDir()
    else:
        
        # if name has not match, directory is created under that name
        os.mkdir(dirName)
        
    # user chooses to go back to main menu or exit program
    menuChoice = input("Would you like to make another selection? 'y' or 'n': \n")
    if menuChoice == 'y':
        operate_menu(menu_choices)
    else:
        print("Session Closed")

# function to create project (work stream, creates folder)
def createProj():
    dirChoice = input("Please choose a directory to save project file. If you're unsure of the name type 'list' to see list of directories: ")
    if dirChoice == 'list':
        
        # list of available directories is printed to terminal, user taken back to choose project folder name
        print("Here is the list of available directories: ", os.listdir())
    
        createProj()
    elif os.path.isdir(dirChoice) == False:
        
        # if directory doesn't exist, user can choose whether they want to create one, if not session closes/taken back to main menu
        choice = input("This directory does not exist, would you like to create one? 'y' or 'n': \n")
        
        if choice == 'n':
            choice2 = input("Would you like to make another selection? 'y' or 'n': \n")
            if choice2 == 'n':
                print("Session Closed")
            elif choice2 == 'y':
                operate_menu(menu_choices)
            else:
                print("Command not recognised. closing program")
        elif choice == 'y':
            createDir()
    else:
        
        # if directory does exist, user is asked to name new project folder
        print(dirChoice, "directory chosen")
        projName = input('Enter project folder name: ')
        
        # if project name file exists, user is directed to choose a different name
        if os.path.isdir(projName) == True:
            print("Project folder already exists, please choose a different name")
            createProj()
        else:
            
            # else, file path of new project folder (which is created) and directory are joined, user asked to make another selection
            file_path = os.path.join(dirChoice, projName)
            os.mkdir(file_path)
            print("Success", projName, "created under", dirChoice)
            choice = input("Would you like to make another selection? 'y' or 'n': \n")
            if choice == 'n':
                print("Session Closed")
            elif choice == 'y':
                operate_menu(menu_choices)
            else:
                print("Command not recognised. closing program")

# function to create a work stage within stream specified (creates stage folder)
def createWkStr():
    projChoice = input("Please choose a project folder name you would like to create a stage under: ")
    folder_path = searchFolder(projChoice)
    
    # if folder path of project specified doesn't exist, user is asked to create one
    if folder_path is None:
        choice = input("Project folder doesn't exist, would you like to create one?")
        if choice == 'n':
            print("Session Closed")
        elif choice == 'y':
            createProj()
        else:
            print("Command not recognised. closing program")
    
    # else, user is asked to create a new stage name
    if folder_path is not None:
        stageName = input("Enter work stage name to add to project folder")
        
        # if stage file path already exists, user is redirected to choose a different name
        if os.path.isfile(stageName) == True:
            print("Stage already exists, please choose another name")
            createWkStr()
        else:
            
            # else, stage folder is created and path is joined to root
            file_path = os.path.join(folder_path, stageName)
            os.mkdir(file_path)

# function to add new meeting notes to project folder
def addNotes():
    projChoice = input("Enter project folder name you would like to add meeting notes to: ")
    folder_path = searchFolder(projChoice)
    
    # if folder path doesn't exist, user can choose to create one
    if folder_path is None:
        choice = input("Project folder doesn't exist, would you like to create one?")
        if choice == 'n':
            print("Session Closed")
        elif choice == 'y':
            createProj()
        else:
            print("Command not recognised. closing program")
    
    # if folder path exists, user enters name of new text file to add to project folder     
    if folder_path is not None:
        file_name = input("Enter meeting file name to add to project folder: \n")
        
        # ensures file name is in format {name}.txt
        new_file_name = file_name + ".txt"
        file_path = os.path.join(folder_path, new_file_name)
        
        # if text file path already exists i.e., name has already been used, user is directed to choose a different name
        if os.path.exists(file_path):
            print("File already exists, please choose a different name")
            addNotes()
        else:
            
            # else, new text file under name provided is created, user is prompted to enter meeting notes, file is saved and closed
            print("Creating file")
            f = open(new_file_name, "w")
            fileNotes = ''
            
            # while user doesn't press space then enter, add input to file, then close file and save
            while fileNotes != ' ':
                fileNotes = input("Write below your meeting notes to be saved to", new_file_name, ". Press SPACE then press ENTER to finished entering notes.")
                f.write(fileNotes)
            f.close()
            print("Success")
            print(file_path)
            
            # user is prompted to make a selection (to main menu or exit program)
            choice = input("Would you like to make another selection? 'y' or 'n'?\n")
            if choice == 'n':
                print("Session Closed")
            elif choice == 'y':
                operate_menu(menu_choices)
            else:
                print("Command not recognised. Closing program")

# function to add existing meeting notes to (as a pre-existing file where file paths are then joined together)
def addExistingNotes():
    projChoice = input("Enter project folder name you would like to add existing meeting notes to: ")
    folder_path = searchFolder(projChoice)
    
    # if folder path to project doesn't exist, user is given a choice to create one
    if folder_path is None:
        choice = input("Project folder doesn't exist, would you like to create one?")
        if choice == 'n':
            print("Session Closed")
        elif choice == 'y':
            createProj()
        else:
            print("Command not recognised. closing program")
    
    # if folder path exists, user is prompted to enter the name of pre-existing text file
    if folder_path is not None:
        file_name = input("Enter existing meeting file name to add to project folder: \n")
        
        # ensures file name is in format, {name}.txt
        existing_file_name = file_name + ".txt"
        existing_path = searchFile(existing_file_name)
        
        # if file path doesn't exist, user is asked if they want to create one
        if existing_path is None:
            choice = input("Meeting file doesn't exist, would you like to create one?")
            if choice == 'n':
                print("Session Closed")
            elif choice == 'y':
                addNotes()
            else:
                print("Command not recognised. closing program")
        
        # if file path does exist, text file path and root are joined together and printed to terminal, user is then asked whether they want to make another selection
        if existing_path is not None:
            file_path = os.path.join(folder_path, existing_file_name)
            print("Success")
            print(file_path)
            choice = input("Would you like to make another selection? 'y' or 'n'?\n")
            if choice == 'n':
                print("Session Closed")
            elif choice == 'y':
                operate_menu(menu_choices)
            else:
                print("Command not recognised. Closing program")

# function to add meeting attendees to files
def addAttendees():
    projChoice = input("Enter folder name where your excel file is saved: \n")
    folder_path = searchFolder(projChoice)
    
    # if folder path of project doesn't exist, user is asked whether they want to create one
    if folder_path is None:
        choice = input("Project folder doesn't exist, would you like to create one?")
        if choice == 'n':
            print("Session Closed")
        elif choice == 'y':
            createProj()
        else:
            print("Command not recognised. closing program")
    
    # if folder path of project exists, user is prompted to enter name of excel file to add meeting attendees to
    if folder_path is not None:
        excel_name = input("Enter name of your Excel file: \n")
        
        # ensures file is in format {name}.xlsx
        new_excel_name = excel_name + ".xlsx"
        excel_path = searchFile(new_excel_name)
        
        # is excel path doesn't exist, user is prompted to create a new file with name specified
        if excel_path is None:
            choice = input("This Excel file doesn't exist. Would you like to create a new file under this name and amend to it: 'y' or 'n'")
            if choice == 'n':
                choice2 = input("Would you like to make another selection? 'y' or 'n'?\n")
                if choice2 == 'n':
                    print("Session Closed")
                elif choice2 == 'y':
                    operate_menu(menu_choices)
                else:
                    print("Command not recognised. Session Closed.")
                
            elif choice == 'y':
                print("Creating file...")
                workbook = xlsxwriter.Workbook(excel_path) # creates excel sheet in 'workbook' in the specific path (includes file name)
                worksheet = workbook.add_worksheet() # method used to add new worksheet to the workbook where workbook is now an object
                worksheet = workbook.active # makes worksheet active (to be able to read and write from and to)
                
                # variable to hold number of attendees being added to Excel file
                numPeople = int(input("Please enter as an integer (e.g. 4) how many attendees you are adding to the Excel file: \n"))
                
                # new list to add attendee names which will be added to file
                toWrite = []
                
                # for loop to append to list in range between 0 and number specified in variable numPeople
                for i in range(numPeople):
                    name = input("Enter attendee name number:")
                    toWrite.append(name)
                
                # variable called format which takes the list 'toWrite' and reformats it using pandas function 'Series' in Excel format (i.e., 2D  array)
                format = pd.Series(toWrite)
                
                # prints variable format to show user what is being written to file
                print(format)
                
                # format is written to Excel file and saved, user is prompted to make another selection
                format.to_excel(excel_path,sheet_name='Sheet1')
                workbook.close()
                
                print("Success")
                choice = input("Would you like to make another selection? 'y' or 'n'?\n")
                if choice == 'n':
                    print("Session Closed")
                elif choice == 'y':
                    operate_menu(menu_choices)
                else:
                    print("Command not recognised. Session Closed.")
            
            else:
                print("Command not recognised. Session Closed.")
        
        
        if excel_path is not None:
            # variable to hold number of attendees being added to Excel file
            numPeople = int(input("Please enter as an integer (e.g. 4) how many attendees you are adding to the Excel file: \n"))
            
            # new list to add attendee names which will be added to file
            toWrite = []
            
            # for loop to append to list in range between 0 and number specified in variable numPeople
            for i in range(numPeople):
                name = input("Enter attendee name number:")
                toWrite.append(name)
                
            # variable called format which takes the list 'toWrite' and reformats it using padnas function 'Series' in Excel format (i.e., 2D array)
            format = pd.Series(toWrite)
            
            # prints variable format to show user what is being written to file
            print(format)
            
            # format is written to Excel file and saved, user is prompted to make another selection
            format.to_excel(excel_path,sheet_name='Sheet1')
            workbook.close()
            
            print("Success")
            choice = input("Would you like to make another selection? 'y' or 'n'?\n")
            if choice == 'n':
                print("Session Closed")
            elif choice == 'y':
                operate_menu(menu_choices)
            else:
                print("Command not recognised. Session Closed.")
    
# function to add pre-existing files to workstream folder 
def addWkStrFiles():
    folderChoice = input("Enter name of workstream (project) folder to add your file to:  \n")
    folderPath = searchFolder(folderChoice)
    
    # if project path doesn't exist, user is prompted to make one'
    if folderPath is None:
        choice = input("Workstream doesn't exist, would you like to create one? 'y' or 'n': \n")
        if choice == 'n':
            print("Exiting to Main Menu...")
            operate_menu(menu_choices)
            
        if choice == 'y':
            createProj()
    
    # if project path does exist then user is prompted to type file extension (type) and name of file they want to select
    if folderPath is not None:
        typeChoice = input("Enter file extension e.g. '.xlsx', '.txt', '.csv': \n")
        fileChoice = input("Enter file name to add to workstream: \n")
        
        # ensures file is in correct format to be searched with
        filePath = searchFile(fileChoice + typeChoice)
        
        # if file path doesn't exist, user is returned to main menu
        if filePath is None:
            print("File doesn't existing. Exiting to Main Menu...")
            operate_menu(menu_choices)
        
        # if file path does exist, file path and project path are joined, user is then prompted to make another selection
        if filePath is not None:
            print("Adding file to folder...")
            os.path.join(folderPath, (fileChoice + typeChoice))
            print("Success!")
            choice = input("Would you like to make another selection? 'y' or 'n'?\n")
            if choice == 'n':
                print("Session Closed")
            elif choice == 'y':
                operate_menu(menu_choices)
            else:
                print("Command not recognised. Session Closed.")
            
# function to add participants to project Excel file
def addParticipants():
    file_name = input("Name new Project Participant Excel file: \n")
    
    # ensures file is in correct format: {name}.xlsx
    file_Path = searchFile(file_name + 'xlsx')
    
    # if file path does already exist, user is prompted to choose a different name
    if file_Path is not None:
        print("That file already exists, please choose a different name\n")
        addParticipants()
    
    # if file path does not exist, a new workbook is created to add project participants
    if file_Path is None:
        workbook = xlsxwriter.Workbook(file_Path) # creates excel sheet in 'workbook' in the specific path (includes file name)
        worksheet = workbook.add_worksheet() # method used to add new worksheet to the workbook as workbook is now an object
        worksheet = workbook.active # makes worksheet active (to be able to read and write from and to)
        
        # variable to hold number of participants being added to Excel file
        numPeople = int(input("Please enter as an integer (e.g. 4) how many participants you are adding to the Excel file: \n"))
        
        # new list to add participant names which will be added to file
        toWrite = []
        
        # for loop to append to list in range between 0 and number specified in variable numPeople
        for i in range(numPeople):
            name = input("Enter attendee name number:")
            toWrite.append(name)
        
        # variable called format which takes the list 'toWrite' and reformats it using pandas function 'Series' in Excel format
        format = pd.Series(toWrite)
        
        # prints variable format to show user what is being written to file
        print(format)
        
        # format is written to Excel file and saved, user is prompted to make another selection
        format.to_excel(file_Path,sheet_name='Sheet1') 
        workbook.close()
        
        choice = input("Would you like to make another selection? 'y' or 'n'?\n")
        if choice == 'n':
            print("Session Closed")
        elif choice == 'y':
            operate_menu(menu_choices)
        else:
            print("Command not recognised. Session Closed.")   

# function to load directories within current working directories (shows all folders)
def ldDir():
    print("Here is a list of folders within: ",os.getcwd())
    print(os.listdir())

# function to load project participants from CSV file
def ldProjParticipants():
    fileChoice = input("Enter name of CSV file you wish to load project participant names from: \n")
    
    # ensures correct file format: {name}.csv
    file_path = searchFile(fileChoice + ".csv")
    
    # if file path does not exist, user is prompted to make a selection directing them either to main menu or to add participants
    if file_path is None:
        choice = input("That file doesn't exist, would you like to create one?: 'y' or 'n': \n")
        if choice == 'n':
            choice = input("Would you like to make another selection? 'y' or 'n'?\n")
            if choice == 'n':
                print("Session Closed")
            elif choice == 'y':
                operate_menu(menu_choices)
            else:
                print("Command not recognised. Session Closed.")
            
        elif choice == 'y':
            addParticipants()
    
        else:
            print("Command not recognised. Session CLosed.")
    
    # if file path does exist, the CSV file is opened in read mode
    if file_path is not None:
        csvFile = open(fileChoice + ".csv", 'r')
        
        # CSV file is split by new lines '\n' into rows 
        rows = csvFile.read().split('\n')
        
        # a new dictionary is created
        person_dict = {}
        
        # rows are split further into headers by ','
        header = rows[0].split(',')
        
        # project index is classified as header named: 'Project Name'
        project_index = header.index('Project Name')
        
        # participant index is classified as header named: 'Project Participants'
        participants_index = header.index('Project Participants')
        
        # for each row in row index 1, each cell containing project name which is not contained in the dictionary person_dict, it is added to the dictionary and appends project participant
        for row in rows[1]:
            cells = row.split(',')
            project_name = cells[project_index]
            participant_name = cells[participants_index]
            if project_name not in person_dict:
                person_dict[project_name] = []
                
            person_dict[project_name].append(participant_name)
        
        # prints dictionary of project name with allocated project participants
        print(person_dict)
        
        choice = input("Would you like to make another selection? 'y' or 'n'?\n")
        if choice == 'n':
            print("Session Closed")
        elif choice == 'y':
            operate_menu(menu_choices)
        else:
            print("Command not recognised. Session Closed.")

# function to load work streams
def ldWkStr():
    projChoice = input("Enter the name of work stage you wish to load: \n")
    proj_path = searchFolder(projChoice)
    
    # if project folder path doesn't exist, user is prompted to make one
    if proj_path is None:
        choice = input("That project doesn't exist, would you like to create one?: 'y' or 'n': \n")
        if choice == 'n':
            choice = input("Would you like to make another selection? 'y' or 'n'?\n")
            if choice == 'n':
                print("Session Closed")
            elif choice == 'y':
                operate_menu(menu_choices)
            else:
                print("Command not recognised. Session Closed.")
            
        elif choice == 'y':
            createProj()
    
        else:
            print("Command not recognised. Session CLosed.")
    
    # if project folder path does exist, all files and folders within chosen work stream are displayed to terminal, user is prompted to make another selection      
    if proj_path is not None:
        print("Displaying files within:",projChoice)
        os.chdir(proj_path)
        print(os.listdir())
        choice = input("Would you like to make another selection? 'y' or 'n'?\n")
        if choice == 'n':
            print("Session Closed")
        elif choice == 'y':
            operate_menu(menu_choices)
        else:
            print("Command not recognised. Session Closed.")

# function to load work stages
def ldStage():
    stageChoice = input("Enter the name of work stage you wish to load: \n")
    stage_path = searchFolder(stageChoice)
    
    # if folder path doesn't exist, user is prompted to make one
    if stage_path is None:
        choice = input("That stage doesn't exist, would you like to create one?: 'y' or 'n': \n")
        if choice == 'n':
            choice = input("Would you like to make another selection? 'y' or 'n'?\n")
            if choice == 'n':
                print("Session Closed")
            elif choice == 'y':
                operate_menu(menu_choices)
            else:
                print("Command not recognised. Session Closed.")
            
        elif choice == 'y':
            createWkStr()
    
        else:
            print("Command not recognised. Session CLosed.")
    
    # if folder path does exist, every file and folder within specified work stage folder is displayed to terminal      
    if stage_path is not None:
        print("Displaying files within:",stageChoice)
        os.chdir(stage_path)
        print(os.listdir())
        
        # opens CSV file in read mode
        csvFile = open(stageChoice + ".csv", 'r')
        
        # split by new lines '\n' into rows
        rows = csvFile.read().split('\n')
        
        # new dictionary is created
        stage_dict = {}
        
        # rows split furhter into headers by ','
        header = rows[0].split(',')
        
        # stage name index is classified as header named: 'Stage Name'
        stage_name_index = header.index('Stage Name')
        
        # status index is classified as header named: 'Stage Status'
        status_index = header.index('Stage Status')
        
        # for each row in row index 1, each cell containing stage name which is not contained in the dictionary stage_dict, it is added to the dictionary and appends stage status
        for row in rows[1]:
            cells = row.split(',')
            stage_name = cells[stage_name_index]
            status = cells[status_index]
            if stage_name not in stage_dict:
                stage_dict[stage_name] = []
                
            stage_dict[stage_name].append(status)
        
        # prints dictionary of stage names and status of each
        print(stage_dict)
        
        choice = input("Would you like to make another selection? 'y' or 'n'?\n")
        if choice == 'n':
            print("Session Closed")
        elif choice == 'y':
            operate_menu(menu_choices)
        else:
            print("Command not recognised. Session Closed.")
    
# function to load existing meeting notes file
def ldMeetingNotes():
    file_name = input("Enter name of file with meeting notes you wish to load: \n")
    
    # ensures file is in correct format: {name}.xlsx
    new_file_name = file_name + ".xlsx"
    file_path = searchFile(new_file_name)
    
    # if file path doesn't exist, user is prompted to create one
    if file_path is None:
        choice = input("That file doesn't exist, would you like to create one? 'y' or 'n': \n")
        
        if choice == 'n':
            choice = input("Would you like to make another selection? 'y' or 'n'?\n")
            if choice == 'n':
                print("Session Closed")
            elif choice == 'y':
                operate_menu(menu_choices)
            else:
                print("Command not recognised. Session Closed.")
            
        elif choice == 'y':
            addNotes()
    
        else:
            print("Command not recognised. Session CLosed.")
    
    # if file path does exist, file name is read using pandas module read_excel into variable 'content'
    if file_path is not None:
        content = pd.read_excel(new_file_name)
        print("Success")
        
        # content of Excel file is printed to terminal, user is prompted to make selection
        print(content)
        choice = input("Would you like to make another selection? 'y' or 'n'?\n")
        if choice == 'n':
            print("Session Closed")
        elif choice == 'y':
            operate_menu(menu_choices)
        else:
            print("Command not recognised. Session Closed.")
        
# function to load excel file containing meeting attendees
def ldMeetingAttendees():
    excel_name = input("Enter name of Excel file with meeting attendees you wish to load: \n")
    
    # ensures file is in correct format: {name}.xlsx
    new_excel_name = excel_name + ".xlsx"
    excel_path = searchFile(new_excel_name)
    
    # if folder path does not exist, user is given option to create one
    if excel_path is None:
        choice = input("That Excel file doesn't exist, would you like to create one? 'y' or 'n': \n")
        
        if choice == 'n':
            choice = input("Would you like to make another selection? 'y' or 'n'?\n")
            if choice == 'n':
                print("Session Closed")
            elif choice == 'y':
                operate_menu(menu_choices)
            else:
                print("Command not recognised. Session Closed.")
            
        elif choice == 'y':
            addAttendees()
    
        else:
            print("Command not recognised. Session CLosed.")
            
    # if file path does exist, content of file is read by pandas module read_excel and printed to terminal, user is prompted to make another decision
    if excel_path is not None:
        content = pd.read_excel(new_excel_name)
        print("Success")
        print(content)
        choice = input("Would you like to make another selection? 'y' or 'n'?\n")
        if choice == 'n':
            print("Session Closed")
        elif choice == 'y':
            operate_menu(menu_choices)
        else:
            print("Command not recognised. Session Closed.")

# function to delete directory 
def DelDir():
    dirChoice = input("Enter directory folder name to delete: \n")
    
    # function to check if directory exists is called
    dirPath = searchFolder(dirChoice)
    
    # if directory path doesn't exist, user is taken back to main menu
    if dirPath is None:
        print("Directory doesn't exist. Exiting to Main Menu...")
        operate_menu(menu_choices)
    
    # if directory path does exist, user is warned of consequences of action, if they continue, directory and all contents are deleted, if not user is taken back to main menu, else session is closed
    if dirPath is not None:
        choice = input("Deleting this directory will delete all subfolders and files. This is not reversible. Would you like to continue? 'y' or 'n': \n")
        if choice == 'n':
            print("Exiting to Main Menu...")
            operate_menu(menu_choices)
            
        elif choice == 'y':
            delNonEmptyDir(dirPath)
        
        else:
            print("Command not recognised. Session Closed.")

# function to delete project file
def DelProj():
    projChoice = input("Enter project folder name to delete: \n")
    # function to check if project folder path exists is called
    projPath = searchFolder(projChoice)
    
    # if folder path doesn't exist, user is taken back to main menu
    if projPath is None:
        print("Project folder doesn't exist. Exiting to Main Menu...")
        operate_menu(menu_choices)
    
    # if folder path does exist, user is warned of consequences of action, if they continue, folder and subfolders and files are deleted, if not user is taken back to main menu, else session is closed
    if projPath is not None:
        choice = input("Deleting this project will delete all subfolders and files. This is not reversible. Would you like to continue? 'y' or 'n': \n")
        if choice == 'n':
            print("Exiting to Main Menu...")
            operate_menu(menu_choices)
            
        elif choice == 'y':
            delNonEmptyDir(projPath)
        
        else:
            print("Command not recognised. Session Closed.")

# function to delete folder representing a stage in a project
def DelStage():
    projChoice = input("Enter stage folder name to delete: \n")
    # function to check if folder path exists is called
    projPath = searchFolder(projChoice)
    
    # if folder path does not exist, user is taken back to main menu
    if projPath is None:
        print("Stage folder doesn't exist. Exiting to Main Menu...")
        operate_menu(menu_choices)
    
    # if folder path does exist, user is warned of consequences of action, if they continue, folder with subfolders and files are deleted, if not they are taken back to main menu, else session closes
    if projPath is not None:
        choice = input("Deleting this stage will delete all subfolders and files. This is not reversible. Would you like to continue? 'y' or 'n': \n")
        if choice == 'n':
            print("Exiting to Main Menu...")
            operate_menu(menu_choices)
            
        elif choice == 'y':
            delNonEmptyDir(projPath)
        
        else:
            print("Command not recognised. Session Closed.")

# funtion to delete notes (text files i.e., meeting notes) from directory/folder 
def DelNotes():
    notesChoice = input("Enter notes file name to delete: \n")
    # function to check if file path exsits is called
    notesPath = searchFolder(notesChoice)
    
    # if file does not exist, user is taken back to main menu
    if notesPath is None:
        print("File doesn't exist. Exiting to Main Menu...")
        operate_menu(menu_choices)
    
    # if file path does exist, user is warned about consequences of action, if they continue, file is deleted, if not they are returned to main menu, else session closes
    if notesPath is not None:
        choice = input("Deleting this file is not reversible. Would you like to continue? 'y' or 'n': \n")
        if choice == 'n':
            print("Exiting to Main Menu...")
            operate_menu(menu_choices)
            
        elif choice == 'y':
            print("Deleting file...")
            os.remove(notesPath)
            print("Success!")
        
        else:
            print("Command not recognised. Session Closed.")

# function to exit menu and terminate program
def exitProgram():
    print("Session Closed")

# dictionary with menu values directing to corresponding functions available for user to choose from (when picked the function chosen will be called)
menu_choices = {
    "1": createDir,
    "2": createProj,
    "3": createWkStr,
    "4": addNotes,
    "5": addExistingNotes,
    "6": addAttendees,
    "7": addWkStrFiles,
    "8": addParticipants,
    "A": ldDir,
    "B": ldProjParticipants,
    "C": ldWkStr,
    "D": ldStage,
    "E": ldMeetingNotes,
    "F": ldMeetingAttendees,
    "G": DelDir,
    "H": DelProj,
    "I": DelStage,
    "J": DelNotes,
    "n": exitProgram 
}

# main menu operator function > shows user list of functions to choose from
def operate_menu(menuList):
    # setting current working directory as '/Users' (makes it easier to search for files)
    os.chdir('/Users')
    cwd = str(os.getcwd())
    
    # main menu set of selection choices displayed to user is terminal
    print("You are currently working in: " + cwd)
    print("Please choose a menu subsection from the list below: ")
    menuChoice = input("----------Create Section----------\n1. Create Directory\n2. Create Workstream (Project)\n3. Create Work Stage\n4. Create New Meeting Notes\n----------Add Section----------\n5. Add Existing Meeting Notes\n6. Add Meeting Attendees\n7. Add to Workstream Files\n----------Load Section----------\nA. Load Directory\nB. Load Project Participants Information\nC. Load Workstreams & Info\nD. Load Stages & Info\nE. Load Meeting Notes\nF. Load Meeting Attendees\n----------Delete Section----------\nG. Delete Directory\nH. Delete Project\nI. Delete Stage\nJ. Delete Meeting Notes\nENTER 'n' TO EXIT PROGRAM\n")
    
    # if the menu choice selected exists in the menu dictionary, that function is then called, else program terminates
    if menuChoice in menuList:
        print("Success")
        menuList[menuChoice]()
    else:
        print("Attempt Unsuccessful")

# calls operate_menu function (main program)
def main():
    operate_menu(menu_choices)

if __name__ == "__main__":
    main()

# Notes (important functions for me to remember)
# os.chdir() > change current working directory
# os.chroot() > change root directory of current process to specific path
# os.getcwd() > get current working directory > current working directory is folder in which python script is operating
# os.listdir() > returns list of names of the entries in a directory
# os.makedirs() > creates a directory recursively
# os.mkdir() > makes directory with specific mode
# .open() > opens existing file
# os.remove() > delete file
# os.rmdir() > delete empty directory
